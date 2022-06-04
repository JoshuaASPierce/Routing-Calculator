#Joshua Pierce, ID: 001212293
#C950
#Imports here:
from openpyxl.reader.excel import load_workbook
import math



#assumptions:
#Loading and delivery are instantaneous
ats = 18.0 #Average truck speed in miles per hour
maxTruckPack = 16 #Maximum packages per truck
maxTrucks = 2 #Number of trucks
maxDrivers = 2 #Number of drivers available at a time
startTime = 8 #Earliest time a driver can leave the hub
maxDistance = 140.0 #Maximum amount of mileage that both trucks can travel

#Spreadsheet info:
firstRow = 9 #First row that has data on it in the package file
lastRow = 48 #last row with data on it in the package file


packageTupleList = [] #A list of tuples as described above
totalMiles = 0
hubPackageList = [] #A list of package class objects that havent been loaded to a truck yet.
hubTruckList = [] # A list of all truck class objects
addressList = [] # a list of all address class objects
priorityOne = [] #Priority one packages
priorityTwo = [] #priority two packages
priorityThree = [] #priority three packages, AKA the late arrivals
packageHashList = [] #The hashlist of all packages


class Package:
    def __init__(self, id, address, zip, deadline, mass, special):
        self.id = id #Package id as per package file
        self.address = address #package address
        self.zip = zip #package zipcode
        self.deadline = deadline #the latest at which the package can be delivered
        self.mass = mass #package weight in Kilograms
        self.special = special #special notes, either specific truck needed, delay to start, wrong address, or must be delivered with other packages

class Address:
    def __init__(self, id, address, distanceDict, priority):
        self.id = id #id given at construction
        self.address = address #address per distance table
        self.distanceDict = distanceDict #A list of tuples, each tuple has an address ID and a distance
        self.priority = priority

class Van:
    def __init__(self, id, packageList):
        self.id = id
        self.distance = 0 #Total driven by truck
        self.packageList = packageList #The index's of the lists this truck has

#This function resets all lists and data values to their default values.
def dataReset():
    hubPackageList.clear() #A list of package class objects that havent been loaded to a truck yet.
    hubTruckList.clear() # A list of all truck class objects
    addressList.clear() # a list of all address class objects
    priorityOne.clear() #Priority one packages
    priorityTwo.clear() #priority two packages
    priorityThree.clear()
    packageHashList.clear() #The hashlist of all packages
    packageTupleList.clear()

#Inserts a hash of the package with info given
def insertPackage(id, address, deadline, city, zip, weight, status): 
    #First check for any hashes that have a matching id
    for each in packageHashList:
        if each['id'] == id:
            return False
    packageHashList.append({'id': id, 'address': address, 'deadline': deadline, 'city': city, 'zip': zip, 'weight': weight, 'status': status})

#updates the given package status based on the id
def updatePackage(id, status):
    for each in packageHashList:
        if each['id'] == id:
            each['status'] = status

#this function returns a list of packages from the hash that meet the given criteria
def returnPackage(lookupComp, comp): #lookupComp is the thing you are looking up by, eg, 'id', 'address', or 'deadline'. comp is the actual lookup info, eg '1', '2100 asddd', or '10:30'
    tempList = []
    if lookupComp == 'id':
        comp = int(comp)
    if lookupComp == 'status':
        if comp == "Delivered":
            for each in packageHashList:
                if comp in each[lookupComp]:
                    tempList.append(each)
            return tempList
    for each in packageHashList:
        if each[lookupComp] == comp:
            tempList.append(each)
    return tempList

#Input a list of addresses, return the list, sorted
def dynamicProgrammingApproach(inList):
    if len(inList) < 2:
        return inList
    helpList = []
    for each in inList:
        helpList.append([each])
    inList = helpList
    while len(inList) > 1:
        
        w = '' #where the matching stuff is. if 'ff', then first, first, and so on
        x = 100 #Distance
        y = 1 #first list position in inList
        z = 1 #Second list position in inList
        for each in inList:
            for every in inList:
                if each != every:
                    if each[0].distanceDict[every[0].id] < x:
                        w = 'ff'
                        x = each[0].distanceDict[every[0].id]
                        y = each
                        z = every
                    if each[0].distanceDict[every[-1].id] < x:
                        w = 'fl'
                        x = each[0].distanceDict[every[-1].id]
                        y = each
                        z = every
                    if each[-1].distanceDict[every[-1].id] < x:
                        w = 'll'
                        x = each[-1].distanceDict[every[-1].id]
                        y = each
                        z = every
                    if each[-1].distanceDict[every[0].id] < x:
                        w = 'lf'
                        x = each[-1].distanceDict[every[0].id]
                        y = each
                        z = every
        if w == 'ff':
            inList.append(listAppendFF(y, z))
            inList.remove(y)
            inList.remove(z)
        elif w == 'fl':
            inList.append(listAppendFB(y, z))
            inList.remove(y)
            inList.remove(z)
        elif w == 'll':
            inList.append(listAppendBB(y, z))
            inList.remove(y)
            inList.remove(z)
        elif w == 'lf':
            inList.append(listAppendBF(y, z))
            inList.remove(y)
            inList.remove(z)
    return inList[0]

def listAppendFF(list1, list2): #Given two lists, returns them as list2 reversed and added to the front of list 1
    list2.reverse()
    list2.extend(list1)
    return list2

def listAppendFB(list1, list2): #Given two lists, returns them as list2 appended to front of list1
    list2.extend(list1)
    return list2

def listAppendBF(list1, list2): #Given two lists, returns them as list2 appended to back of list1
    list1.extend(list2)
    return list1

def listAppendBB(list1, list2): #Given two lists, returns them as list2 backwards, appended to back of list1
    list2.reverse()
    list1.extend(list2)
    return list1

#Input an id, returns the address class
def addressIdFind(addressId):
    for x in addressList:
        if x.id == addressId:
            return x

#Input a string of the address, returns the address class
def addressFind(addressIn):
    for x in addressList:
        if x.address == addressIn:
            return x

#Loads the addresses from the spreadsheet given
def load_Addresses(spreadsheet):
    #Creates address class objects based on the excel spreadsheet
    #load the spreadsheet
    workbook = load_workbook(filename=spreadsheet + ".xlsx")
    sheet = workbook.active
    i = 1
    #for each adress on column B, create a new Address class object with a unique id
    for row in sheet.iter_rows(min_row=9, max_row=35, min_col=2, max_col=2): 
        addressList.append(Address(i, ''.join(filter(str.isalnum, str(row[0].value))), {}, 0)) 
        i += 1
    i = 0
    #For each new address, go down the rows and add tuples to the distanceDict based on the id and distance
    #Created as dictionaries, where the key is the address ID, and the value is the distance
    for row in sheet.iter_rows(min_row=9, max_row=35, min_col=3, max_col=29):
        r = 1
        for each in row:
            if type(each.value) is int or type(each.value) is float:
                addressList[i].distanceDict.update({r : each.value})
            r += 1
                
        i += 1
    i = 0
    for col in sheet.iter_cols(min_row=9, max_row=35, min_col=3, max_col=29):
        r = 1
        for each in col:
            if type(each.value) is int or type(each.value) is float:
                addressList[i].distanceDict.update({r : each.value})
            r += 1
        i += 1
    
    helpList = dynamicProgrammingApproach(addressList)
    i = 1
    for address in helpList:
        address.priority = i
        i += 1

#Loads the packages from the spreadsheet given
def load_Packages(spreadsheet):
    #Creates class objects based on the excel spreadsheet
    #load the excel spreadsheet
    workbook = load_workbook(filename=spreadsheet + ".xlsx")
    sheet = workbook.active
    #For each line, create a package class object based on info in the cells, then append to the hubPackageList
    for row in sheet.iter_rows(min_row=firstRow, max_row=lastRow, min_col=1, max_col=8):
        hubPackageList.append(Package(int(row[0].value), addressFind(''.join(filter(str.isalnum, str(row[1].value + str(row[4].value))))) , str(row[4].value), str(row[5].value), int(row[6].value), str(row[7].value)))
        insertPackage(int(row[0].value), str(row[1].value), str(row[5].value), str(row[2].value), str(row[4].value), int(row[6].value), 'at the hub')
    hubPackageList.sort(key=lambda x: x.address.priority)
    for package in hubPackageList:
        if (package.deadline != "EOD") and (package.special.find("Delayed on flight") == -1): #if the package is supposed to be delivered before end of day, put it in the priority queue
            priorityOne.append(package)
        elif (len(package.special) != 4) and (package.special.find("Wrong address listed") == -1) and (package.special.find("Delayed on flight") == -1): 
            priorityOne.append(package) #if the package has special instructions not listed as "Wrong address", put it in priority queue
        elif (package.special.find("Delayed on flight") != -1) or (package.special.find("Wrong address listed") != -1):
            priorityThree.append(package)
        else:
            priorityTwo.append(package) #if the package has no special instructions and is to be delivered by end of day, put in priority two queue
    
    priorityOne.sort(key=lambda x: x.address.priority)
    priorityTwo.sort(key=lambda x: x.address.priority)
    priorityThree.sort(key=lambda x: x.address.priority)

#Converts miles to time as a string
def milesToTime(miles): #Put in miles, get back current time as a string
    minutes = int((miles/ats) * 60)
    hours = int((minutes / 60) + startTime)
    minutes = int(minutes % 60)
    if minutes < 10:
        return str(hours) + ":0" + str(minutes)
    else:
        return  str(hours) + ":" + str(minutes)

#converts time to miles traveled
def timeToMiles(hours, minutes): #put in the current time as hours, minutes, return miles as a float
    minutes += hours*60 #total minutes
    x = ats/60 #miles per minute
    return float(minutes) * x

#This distributes the packages amongst the lists
def distributePackages():

    i = 0
    arrayNum = int(math.ceil(len(hubPackageList)/maxTruckPack))
    global totalMiles
    tempOne = priorityOne
    tempTwo = priorityTwo
    tempThree = priorityThree

    for x in range(arrayNum): #Create a number of lists inside packageTupleList equal to the number needed to hold all necessary packages
        packageTupleList.append([])
    
    for x in range(arrayNum):
        while (len(packageTupleList[x]) <= maxTruckPack) and (tempOne or tempTwo or tempThree):
            if tempOne:
                if packageTupleList[x]:
                    packageTupleList[x].append([tempOne[0]]) #put the package in the list
                    tempOne.pop(0)
                    packageTupleList[x][-1].append(packageTupleList[x][-1][0].address.distanceDict[packageTupleList[x][-2][0].address.id])#add the distance between the package and the one before it
                else:
                    packageTupleList[x].append([tempOne[0]]) #put the package in the list
                    tempOne.pop(0)
                    packageTupleList[x][-1].append(packageTupleList[x][-1][0].address.distanceDict[1])
            elif tempTwo:
                if packageTupleList[x]:
                    packageTupleList[x].append([tempTwo[0]]) #put the package in the list
                    tempTwo.pop(0)
                    packageTupleList[x][-1].append(packageTupleList[x][-1][0].address.distanceDict[packageTupleList[x][-2][0].address.id])#add the distance between the package and the one before it
                else:
                    packageTupleList[x].append([tempTwo[0]]) #put the package in the list
                    tempTwo.pop(0)
                    packageTupleList[x][-1].append(packageTupleList[x][-1][0].address.distanceDict[1])
            else:
                if packageTupleList[x]:
                    packageTupleList[x].append([tempThree[0]]) #put the package in the list
                    tempThree.pop(0)
                    packageTupleList[x][-1].append(packageTupleList[x][-1][0].address.distanceDict[packageTupleList[x][-2][0].address.id])#add the distance between the package and the one before it
                else:
                    packageTupleList[x].append([tempThree[0]]) #put the package in the list
                    tempThree.pop(0)
                    packageTupleList[x][-1].append(packageTupleList[x][-1][0].address.distanceDict[1])
        packageTupleList[x].append([addressIdFind(1), packageTupleList[x][-1][0].address.distanceDict[1]])
    for each in packageTupleList:
        distance = 0
        for every in each:
            totalMiles += every[1]
            distance += every[1]
            every.append(distance)
    
#This finds the truck with the shortest final list
def findShortestTruck():
    x = maxDistance
    y = Van
    for each in hubTruckList:
        if packageTupleList[each.packageList[-1]][-1][-1] < x:
            x = packageTupleList[each.packageList[-1]][-1][-1]
            y = each
    return y

#This function will assign lists of packages to trucks     
def givePackageToTrucks():
    for each in range(maxTrucks): #Assign initial lists based on number of trucks
        hubTruckList.append(Van(each, [each]))
    #Find the truck with the shortest list, and assign it the last list
    findShortestTruck().packageList.append(2)

#this function delivers all packages up to a certain distance
def deliverPackagesOnRoute(maxDistance):
    #For each truck, go down their packageList until they are at max distance, and stop
    for each in hubTruckList: #for each truck
        for every in each.packageList: #For each list
            for x in packageTupleList[every]: #for each item in that list
                each.distance += x[1] #Add the travel distance
                if each.distance <= maxDistance: #If the distance traveled by this truck is less than the max distance for the function, deliver this package
                    updatePackage(x[0].id, 'delivered at: ' + milesToTime(each.distance))
                elif "delivered" in returnPackage("id" ,packageTupleList[every][0][0].id)[0]["status"]: #If the first package is delivered
                    #For each package in that list that isnt delivered, put it as 'en route'
                    for y in packageTupleList[every]:
                        if "at the hub" in returnPackage("id", y[0].id)[0]["status"]:
                            updatePackage(y[0].id, "en route")

#The primary GUI code along with all the code needed to actually run the program.
def main():
    print("Hello and welcome to Packa.os")
    print("Now Loading")
    load_Addresses("WGUPS Distance Table")
    load_Packages("WGUPS Package File")
    distributePackages()
    givePackageToTrucks()
    print("Done")
    currentMiles = 0.0
    while True:
        
        print("It is currently: " + milesToTime(currentMiles))
        print("Please make a selection:\n1: Lookup Package\t3: Change Time\n2: Print All Packages\t4: Quit program")
        currentSelect = input()
        if currentSelect not in ["1","2","3","4"]:
            print("Invalid Input")
            continue
        elif currentSelect == "4":
            print("Goodbye")
            break
        elif currentSelect == "1":
            print("You have chosen to lookup a package.\n please choose a lookup option:\n1: ID\t5: Zip Code\n2: Address\t6: Weight\n3: Deadline\t7: Status\n4: city")
            lookupSelect = input()
            if lookupSelect not in ["1","2","3","4","5","6","7"]:
                print("Invalid Input")
                continue
            elif lookupSelect == "1":
                lookupInfo = input("\nPlease enter the ID:")
                print(returnPackage("id", lookupInfo))
                x = input("\nPress enter to continue")
                continue
            elif lookupSelect == "2":
                lookupInfo = input("\nPlease enter the Address:")
                print(returnPackage("address", lookupInfo))
                x = input("\nPress enter to continue")
                continue
            elif lookupSelect == "3":
                lookupInfo = input("\nPlease enter the Deadline:")
                print(returnPackage("deadline", lookupInfo))
                x = input("\nPress enter to continue")
                continue
            elif lookupSelect == "4":
                lookupInfo = input("\nPlease enter the City:")
                print(returnPackage("city", lookupInfo))
                x = input("\nPress enter to continue")
                continue
            elif lookupSelect == "5":
                lookupInfo = input("\nPlease enter the Zipcode:")
                print(returnPackage("zip", lookupInfo))
                x = input("\nPress enter to continue")
                continue
            elif lookupSelect == "6":
                lookupInfo = input("Please enter the weight:")
                print(returnPackage("weight", lookupInfo))
                x = input("\nPress enter to continue")
                continue
            elif lookupSelect == "7":
                lookupInfo = input("Please enter the status:")
                print(returnPackage("status", lookupInfo))
                x = input("\nPress enter to continue")
                continue
        elif currentSelect == "2":
            for each in packageHashList:
                print(each)
            x = input("\nPress enter to continue")
            continue
        elif currentSelect == "3":
            print("You have decided to change the current time:")
            hours = input("\nHours past " + str(startTime) + ": ")
            minutes = input("\nMinutes past the hour: ")
            if int(minutes) not in range(0,59):
                print("Invalid input")

                continue
            currentMiles = timeToMiles(float(hours), float(minutes))
            dataReset()
            print("Now Loading")
            load_Addresses("WGUPS Distance Table")
            load_Packages("WGUPS Package File")
            distributePackages()
            givePackageToTrucks()
            deliverPackagesOnRoute(currentMiles)
            print("Done")
            for each in hubTruckList:
                
                print("Truck Number: " + str(each.id) + "   Truck Mileage: " + str(each.distance))
            continue

       
main()

